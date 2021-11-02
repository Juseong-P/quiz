from openpyxl import Workbook
from random import *
wb = Workbook()                                 # 새 워크북 생성
ws = wb.active                                  # 현재 확성화된 sheet 가져옴
ws.title = 'NadoSheet'                          # sheet의 이름 변경
ws1 = wb.create_sheet()                         # 새로운 Sheet 기본 이름으로 생성
ws1.title = 'MySheet'                           # Sheet 이름 변경
ws1.sheet_properties.tabColor = "ff66ff"        # sheet 탭 색상 변경

# sheet 순서 : NadoSheet  MySheet  YourSheet
ws2 = wb.create_sheet('YourSheet')              # 주어진 이름으로 sheet 생성
ws3 = wb.create_sheet('NewSheet', 2)            # 2번째 index에 sheet 생성

new_ws = wb['NewSheet']                         # Dict 형태로 Sheet에 접근
print(wb.sheetnames)                            # 모든 sheet 이름 확인

# Sheet 복사
new_ws['A1'] = "Test"                           # NewSheet A1에 Test 입력
target = wb.copy_worksheet(new_ws)
target.title = 'CopiedSheet'

# cell 값 출력
ws['A1'] = 1
print(ws['A1'])                                 # A1 셀의 정보 출력
print(ws['A1'].value)                           # A1 셀의 '값' 출력
print(ws['A10'].value)                          # 값이 없을 땐 'None'

# row = 1, 2, 3, ...
# column = A(1), B(2), C(3), ...
print(ws.cell(row=1, column=1).value)           # ws['A1'].value
c = ws.cell(column=3, row=1, value=10)          # ws.cell을 이용해서 값을 바로쓰기
print(c.value)

index = 0
for x in range(1, 11):
    for y in range(1, 11):
        # ws.cell(row=x, column=y, value = randint(0,100))
        ws.cell(row=x, column=y, value = index)
        index += 1

# 한 줄씩 데이터 넣기
ws1.append(['번호', '영어', '수학'])
for i in range(1, 11):
    ws1.append([i, randint(0, 100),randint(0, 100)])
# 영어 데이터만 가져오기
# col_B = ws1['B']
# for cell in col_B:
#     print(cell.value)

# for row in tuple(ws1.rows):
    # print(row[1].value)

# col_range = ws1['B:C']                          # 영어, 수학 함께 가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_title = ws1[1]
# for cell in row_title:
#     print(cell.value)

# for col in tuple(ws1.columns):
#     print(col[0].value)

from openpyxl.utils.cell import coordinate_from_string
# row_range = ws1[2:ws1.max_row]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#         print(cell.coordinate, end=' ')
#     print()



wb.save('sample.xlsx')                          # 파일 이름 저장
wb.close()

