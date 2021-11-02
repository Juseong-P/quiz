from openpyxl import load_workbook      # 파일 불러오기
from random import *

wb = load_workbook('sample.xlsx')       # 파일에서 wb을 불러옴
ws = wb['NadoSheet']

# cell 데이터 불러오기
# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row=x, column=y).value, end=' ')
#     print()

# cell 데이터 개수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=' ')
    print()

ws1 = wb['MySheet']
# for row in ws1.iter_rows(min_row=2):
#     if int(row[1].value) > 80:
#         print(row[0].value, "번 학생은 영어 천재")

# 행, 열 삽입
ws1.insert_rows(8)                      # 8번째 줄이 비워짐
ws1.insert_rows(8, 5)                   # 8번째 부터 5줄이 생김
ws1.insert_cols(2)                      # B번 열에 삽입

# 행, 열 삭제
# ws1.delete_rows(8)

# 행, 열 이동
# ws1.move_range("B1:C11", rows=0, cols=1)  # 행 0번, 열 1번 이동
# ws1['B1'].value = '국어'
#
# wb.save('sample_Kor.xlsx')

# 차트 그리기
# from openpyxl.chart import BarChart, Reference
# bar_value = Reference(ws1, min_row=1, max_row=11, min_col=2, max_col=3)
# bar_chart = BarChart()                  # 차트 종류 설정
# bar_chart.add_data(bar_value, titles_from_data=True)           # 차트 데이터 추가
# bar_chart.title = "성적표"
# bar_chart.style = 3
# bar_chart.y_axis.title = "점수"
# bar_chart.x_axis.title = "번호"
# ws1.add_chart(bar_chart, 'E1')          # 차트 넣을 위치 정의
#
# wb.save('sample_chart.xlsx')