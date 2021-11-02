from openpyxl import load_workbook
wb = load_workbook('성적.xlsx')
ws = wb.active

for idx, cell in enumerate(ws['D']):
    if idx == 0:
        continue
    cell.value = 10

ws['H1'] = '총점'
ws['I1'] = '성적'


for x in range(2, ws.max_row + 1):
    ws.cell(row=x, column=8).value = '=SUM(B{}:G{})'.format(x, x)
    total = 0
    for y in range(2, 8):
        total += ws.cell(row=x, column=y).value

    if ws.cell(row=x, column=2).value < 5:
        ws.cell(row=x, column=9).value = 'F'
    elif total >= 90:
        ws.cell(row=x, column=9).value = 'A'
    elif total >= 80:
        ws.cell(row=x, column=9).value = 'B'
    elif total >= 70:
        ws.cell(row=x, column=9).value = 'C'
    else:
        ws.cell(row=x, column=9).value = 'D'

wb.save('성적.xlsx')